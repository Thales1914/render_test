import os
import re
import io
import hashlib
import numpy as np
import pandas as pd
from typing import Optional, Dict, Any
from datetime import datetime, time
from contextlib import contextmanager
from urllib.parse import urlparse

import psycopg2
import psycopg2.extras
from sqlalchemy import create_engine, text as _sqltext

from config import FUSO_HORARIO, HORARIOS_PADRAO, TOLERANCIA_MINUTOS


DATABASE_URL = os.getenv("DATABASE_URL")
if not DATABASE_URL:
    raise RuntimeError("DATABASE_URL não definida.")


DB_URL_SSL = DATABASE_URL if "sslmode=" in DATABASE_URL else f"{DATABASE_URL}?sslmode=require"

engine = create_engine(
    DB_URL_SSL,
    pool_size=10,
    max_overflow=20,
    pool_pre_ping=True,
    pool_recycle=1800,
    future=True,
)

def query_df(sql: str, params: Optional[Dict[str, Any]] = None) -> pd.DataFrame:

    with engine.connect() as conn:
        return pd.read_sql(_sqltext(sql), conn, params=params or {})

_u = urlparse(DATABASE_URL)
DB_HOST, DB_PORT, DB_NAME = _u.hostname, _u.port, _u.path[1:]
DB_USER, DB_PASS = _u.username, _u.password

@contextmanager
def get_db_connection():

    conn_str = (
        f"dbname={DB_NAME} user={DB_USER} password={DB_PASS} "
        f"host={DB_HOST} port={DB_PORT} sslmode=require "
        f"connect_timeout=5 keepalives=1 keepalives_idle=30 "
        f"keepalives_interval=10 keepalives_count=5"
    )
    conn = psycopg2.connect(conn_str, cursor_factory=psycopg2.extras.RealDictCursor)
    try:
        yield conn
    finally:
        conn.close()
        
def get_horario_padrao(filial: int, proximo_evento: str) -> time:
    if filial in (3, 4):
        return time(7, 30) if proximo_evento == "Entrada" else time(17, 30)
    return time(8, 0) if proximo_evento == "Entrada" else time(18, 0)        


def _hash_senha(senha: str) -> str:
    return hashlib.sha256(senha.encode('utf-8')).hexdigest()

def init_db():
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute('CREATE TABLE IF NOT EXISTS empresas (id SERIAL PRIMARY KEY, nome_empresa TEXT NOT NULL UNIQUE, cnpj TEXT)')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS funcionarios (
                    cpf TEXT PRIMARY KEY,
                    codigo TEXT NOT NULL,
                    nome TEXT NOT NULL,
                    senha TEXT NOT NULL,
                    role TEXT NOT NULL,
                    empresa_id INTEGER,
                    cod_tipo TEXT,
                    tipo TEXT,
                    filial TEXT,
                    FOREIGN KEY (empresa_id) REFERENCES empresas (id)
                )
            ''')
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS registros (
                    id TEXT PRIMARY KEY,
                    cpf_funcionario TEXT NOT NULL,
                    nome TEXT NOT NULL,
                    data TEXT NOT NULL,
                    hora TEXT NOT NULL,
                    descricao TEXT NOT NULL,
                    diferenca_min INTEGER NOT NULL,
                    observacao TEXT,
                    FOREIGN KEY (cpf_funcionario) REFERENCES funcionarios (cpf)
                )
            ''')
            
            cursor.execute("SELECT COUNT(*) FROM funcionarios")
            if cursor.fetchone()[0] == 0:
                initial_users = [('admin', 'admin', 'Administrador', _hash_senha('admin123'), 'admin', None, None, None, None)]
                cursor.executemany("INSERT INTO funcionarios (cpf, codigo, nome, senha, role, empresa_id, cod_tipo, tipo, filial) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)", initial_users)
        conn.commit()

def _obter_ou_criar_empresa_id(nome_empresa, cnpj, cursor):
    cursor.execute("SELECT id FROM empresas WHERE lower(nome_empresa) = lower(%s)", (nome_empresa,))
    resultado = cursor.fetchone()
    if resultado:
        cursor.execute("UPDATE empresas SET cnpj = %s WHERE id = %s", (cnpj, resultado[0]))
        return resultado[0]
    else:
        cursor.execute("INSERT INTO empresas (nome_empresa, cnpj) VALUES (%s, %s) RETURNING id", (nome_empresa, cnpj))
        return cursor.fetchone()[0]

def ler_empresas():
    with get_db_connection() as conn:
        return query_df("SELECT id, nome_empresa, cnpj FROM empresas ORDER BY nome_empresa")

def ler_funcionarios_df():
    with get_db_connection() as conn:
        query = "SELECT f.codigo, f.nome, f.cpf, f.cod_tipo, f.tipo, f.filial, f.role, f.empresa_id, e.nome_empresa, e.cnpj FROM funcionarios f LEFT JOIN empresas e ON f.empresa_id = e.id"
        return query_df(query)

def verificar_login(cpf, senha_cod_forte):
    senha_hash = _hash_senha(senha_cod_forte)
    user = None
    with get_db_connection() as conn:
        with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cursor:
            cursor.execute("SELECT * FROM funcionarios WHERE cpf = %s AND senha = %s", (cpf, senha_hash))
            user = cursor.fetchone()
    return (dict(user), None) if user else (None, "CPF ou Senha (Código Forte) inválidos.")

def obter_proximo_evento(cpf):
    hoje_str = datetime.now(FUSO_HORARIO).strftime("%Y-%m-%d")
    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT COUNT(*) AS total FROM registros WHERE cpf_funcionario = %s AND data = %s",
                (cpf, hoje_str)
            )
            result = cursor.fetchone()

            if result is None:
                num_pontos = 0
            elif isinstance(result, dict):
                num_pontos = result.get("total", 0)
            else:
                num_pontos = result[0] if len(result) > 0 else 0

    eventos = list(HORARIOS_PADRAO.keys())
    return eventos[num_pontos] if num_pontos < len(eventos) else "Jornada Finalizada"


def bater_ponto(cpf, nome):
    agora = datetime.now(FUSO_HORARIO)
    proximo_evento = obter_proximo_evento(cpf)
    if proximo_evento == "Jornada Finalizada":
        return "Sua jornada de hoje já foi completamente registada.", "warning"

    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                "SELECT filial FROM funcionarios WHERE cpf = %s",
                (cpf,)
            )
            resultado = cursor.fetchone()
    filial = resultado[0] if resultado else None

    if filial in ("Filial 03", "Filial 3", "Filial 04", "Filial 4"):
        horarios = {
            "Entrada": time(7, 30),
            "Saída":   time(17, 30)
        }
    else:
        horarios = HORARIOS_PADRAO

    hora_prevista     = horarios[proximo_evento]
    datetime_previsto = agora.replace(
        hour=hora_prevista.hour,
        minute=hora_prevista.minute,
        second=0,
        microsecond=0
    )

    diff_bruta = round((agora - datetime_previsto).total_seconds() / 60)
    diff_final = (
        0 if abs(diff_bruta) <= TOLERANCIA_MINUTOS
        else diff_bruta - TOLERANCIA_MINUTOS
        if diff_bruta > 0
        else diff_bruta + TOLERANCIA_MINUTOS
    )

    novo_reg = (
        f"{cpf}-{agora.isoformat()}",
        cpf,
        nome,
        agora.strftime("%Y-%m-%d"),
        agora.strftime("%H:%M:%S"),
        proximo_evento,
        diff_final,
        ""
    )

    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            cursor.execute(
                "INSERT INTO registros (id, cpf_funcionario, nome, data, hora, descricao, diferenca_min, observacao) "
                "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                novo_reg
            )
        conn.commit()

    msg_extra = ""
    if diff_final != 0:
        msg_extra = (
            f" ({diff_final} min de atraso)"
            if diff_final > 0
            else f" ({-diff_final} min de adiantamento)"
        )

    status_final = " (em ponto)"
    if diff_final != 0:
        status_final = ""
    elif diff_bruta != 0:
        status_final = " (dentro da tolerância, registrado como 'em ponto')"

    return (
        f"'{proximo_evento}' registado para {nome} às {agora.strftime('%H:%M:%S')}"
        f"{msg_extra}{status_final}.",
        "success"
    )

def ler_registros_df():
    with get_db_connection() as conn:
        query = "SELECT r.id, f.codigo, r.nome, r.data, r.hora, r.descricao, r.diferenca_min, r.observacao, e.nome_empresa, e.cnpj, f.tipo as setor, f.filial FROM registros r JOIN funcionarios f ON r.cpf_funcionario = f.cpf LEFT JOIN empresas e ON f.empresa_id = e.id"
        df = query_df(query)
    return df.rename(columns={'id': 'ID', 'codigo': 'Código Forte', 'nome': 'Nome', 'data': 'Data', 'hora': 'Hora', 'descricao': 'Descrição', 'diferenca_min': 'Diferença (min)', 'observacao': 'Observação', 'nome_empresa': 'Empresa', 'cnpj': 'CNPJ', 'setor': 'Setor', 'filial': 'Filial'})

def atualizar_registro(id_registro, novo_horario=None, nova_observacao=None):
    try:
        with get_db_connection() as conn:
            with conn.cursor(cursor_factory=psycopg2.extras.DictCursor) as cursor:
                print(f"🛠️ Atualizando registro ID: {id_registro}")

                cursor.execute("SELECT * FROM registros WHERE id = %s", (id_registro,))
                registro_existente = cursor.fetchone()
                if not registro_existente:
                    print("❌ Registro não encontrado")
                    return "Registro não encontrado.", "error"

                campos_atualizados = 0

                if nova_observacao is not None:
                    cursor.execute(
                        "UPDATE registros SET observacao = %s WHERE id = %s",
                        (nova_observacao, id_registro)
                    )
                    campos_atualizados += cursor.rowcount

                if novo_horario is not None:
                    try:
                        novo_obj = datetime.strptime(novo_horario, "%H:%M:%S").time()
                    except ValueError:
                        print("❌ Horário inválido recebido")
                        return "Formato de hora inválido. Use HH:MM:SS.", "error"

                    cursor.execute("""
                        SELECT r.descricao, r.data, r.cpf_funcionario, f.filial
                        FROM registros r
                        JOIN funcionarios f ON f.cpf = r.cpf_funcionario
                        WHERE r.id = %s
                    """, (id_registro,))
                    row = cursor.fetchone()

                    if row:
                        descricao = row['descricao']
                        data_str = row['data']
                        filial_tx = row['filial']

                        filial_num = None
                        if filial_tx:
                            m = re.search(r'\d+', str(filial_tx))
                            filial_num = int(m.group()) if m else None

                        hora_prevista = get_horario_padrao(filial_num, descricao)

                        dt_reg_dia = datetime.strptime(data_str, "%Y-%m-%d")
                        dt_previsto = dt_reg_dia.replace(
                            hour=hora_prevista.hour, minute=hora_prevista.minute, second=0, microsecond=0
                        )
                        dt_novo = dt_reg_dia.replace(
                            hour=novo_obj.hour, minute=novo_obj.minute, second=0, microsecond=0
                        )

                        diff_bruta = round((dt_novo - dt_previsto).total_seconds() / 60)

                        if abs(diff_bruta) <= TOLERANCIA_MINUTOS:
                            diff_final = 0
                        elif diff_bruta > 0:
                            diff_final = diff_bruta - TOLERANCIA_MINUTOS
                        else:
                            diff_final = diff_bruta + TOLERANCIA_MINUTOS

                        cursor.execute(
                            "UPDATE registros SET hora = %s, diferenca_min = %s WHERE id = %s",
                            (novo_horario, diff_final, id_registro)
                        )
                        campos_atualizados += cursor.rowcount

                if campos_atualizados == 0:
                    print("⚠️ Nenhuma alteração feita")
                    return "Nenhuma alteração foi realizada.", "warning"

            conn.commit()
            print("✅ Alterações salvas no banco")

        return "Registro atualizado com sucesso.", "success"

    except psycopg2.Error as e:
        print(f"❌ Erro no banco: {e}")
        return f"Erro no banco de dados: {e}", "error"



def adicionar_funcionario(codigo, nome, nome_empresa, cnpj, cpf, cod_tipo, tipo, filial):
    if not all([codigo, nome, nome_empresa, cpf]):
        return "Campos essenciais (Código Forte, Nome, Empresa, CPF) são obrigatórios.", "error"
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute("SELECT cpf FROM funcionarios WHERE cpf = %s", (cpf,))
                if cursor.fetchone():
                    return f"O CPF '{cpf}' já está em uso.", "warning"
                
                empresa_id = _obter_ou_criar_empresa_id(nome_empresa, cnpj, cursor)
                senha_hash = _hash_senha(codigo)
                cursor.execute(
                    "INSERT INTO funcionarios (cpf, codigo, nome, senha, role, empresa_id, cod_tipo, tipo, filial) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
                    (cpf, codigo, nome, senha_hash, 'employee', empresa_id, cod_tipo, tipo, filial)
                )
            conn.commit()
    except psycopg2.Error as e: return f"Erro no banco de dados: {e}", "error"
    return f"Funcionário '{nome}' adicionado com sucesso!", "success"

def _extrair_filial_do_texto(texto_arquivo):
    texto_lower = texto_arquivo.lower()
    if "matriz" in texto_lower: return "Matriz"
    if "filial 02" in texto_lower or "filial 2" in texto_lower: return "Filial 02"
    if "filial 03" in texto_lower or "filial 3" in texto_lower: return "Filial 03"
    if "filial 04" in texto_lower or "filial 4" in texto_lower: return "Filial 04"
    return "Não Identificada"

def importar_funcionarios_em_massa(df_funcionarios):
    novos_funcionarios, erros, sucesso_count, ignorados_count = [], [], 0, 0
    cpfs_existentes = ler_funcionarios_df()['cpf'].tolist()
    
    colunas_necessarias = ['ARQUIVO', 'EMPRESA', 'CNPJ', 'CODTIPO', 'TIPO', 'CODFORTE', 'NOME', 'CPF']
    if not all(col.upper() in df_funcionarios.columns for col in colunas_necessarias):
        return 0, 0, [f"Erro Crítico: Verifique se as colunas {colunas_necessarias} existem no arquivo."]

    with get_db_connection() as conn:
        with conn.cursor() as cursor:
            empresas_existentes_df = ler_empresas()
            empresas_existentes = dict(zip(empresas_existentes_df['nome_empresa'].str.lower(), empresas_existentes_df['id']))

            for index, row in df_funcionarios.iterrows():
                try:
                    filial = _extrair_filial_do_texto(str(row['ARQUIVO']))
                    nome_empresa = str(row['EMPRESA']).strip()
                    cnpj = str(row['CNPJ']).strip()
                    cod_tipo = str(row['CODTIPO']).strip()
                    tipo = str(row['TIPO']).strip()
                    codigo = str(row['CODFORTE']).strip()
                    nome = str(row['NOME']).strip()
                    cpf_raw = str(row['CPF']).strip()
                    
                    if cpf_raw in cpfs_existentes:
                        ignorados_count += 1
                        continue
                    if not all([codigo, nome, cpf_raw, nome_empresa]):
                        erros.append(f"Linha {index+2}: Dados essenciais (CodForte, Nome, CPF, Empresa) incompletos.")
                        continue
                    
                    empresa_id = empresas_existentes.get(nome_empresa.lower())
                    if not empresa_id:
                        empresa_id = _obter_ou_criar_empresa_id(nome_empresa, cnpj, cursor)
                        empresas_existentes[nome_empresa.lower()] = empresa_id

                    senha_hash = _hash_senha(codigo)
                    novos_funcionarios.append((cpf_raw, codigo, nome, senha_hash, 'employee', empresa_id, cod_tipo, tipo, filial))
                    cpfs_existentes.append(cpf_raw)
                except Exception as e:
                    erros.append(f"Linha {index+2}: Erro - {e}")
            
            if novos_funcionarios:
                try:
                    cursor.executemany("INSERT INTO funcionarios (cpf, codigo, nome, senha, role, empresa_id, cod_tipo, tipo, filial) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)", novos_funcionarios)
                    sucesso_count = len(novos_funcionarios)
                except psycopg2.Error as e:
                    erros.append(f"Erro geral no banco de dados: {e}")
        conn.commit()
        
    return sucesso_count, ignorados_count, erros

def excluir_funcionario(cpf):
    try:
        with get_db_connection() as conn:
            with conn.cursor() as cursor:
                cursor.execute("DELETE FROM registros WHERE cpf_funcionario = %s", (cpf,))
                cursor.execute("DELETE FROM funcionarios WHERE cpf = %s", (cpf,))
            conn.commit()
        return f"Funcionário com CPF {cpf} e todos os seus registros foram excluídos.", "success"
    except psycopg2.Error as e:
        return f"Erro no banco de dados ao excluir funcionário: {e}", "error"

def _formatar_timedelta(td):
    if pd.isnull(td): return "00:00"
    total_seconds = int(td.total_seconds())
    hours, remainder = divmod(total_seconds, 3600)
    minutes, _ = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}"

def gerar_relatorio_organizado_df(df_registros: pd.DataFrame) -> pd.DataFrame:
    if df_registros.empty: return pd.DataFrame()
    df = df_registros.copy()
    df['Descrição'] = df['Descrição'].replace({"Início do Expediente": "Entrada", "Fim do Expediente": "Saída"})
    df_pivot = df.pivot_table(index=['Data', 'Código Forte', 'Nome', 'Empresa', 'CNPJ'], columns='Descrição', values='Hora', aggfunc='first').reset_index()
    df_obs = df.dropna(subset=['Observação']).groupby(['Data', 'Código Forte'])['Observação'].apply(lambda x: ' | '.join(x.unique())).reset_index()
    df_final = pd.merge(df_pivot, df_obs, on=['Data', 'Código Forte'], how='left').fillna({'Observação': ''})
    for evento in ['Entrada', 'Saída']:
        if evento not in df_final.columns: df_final[evento] = np.nan
        df_final[evento] = pd.to_datetime(df_final[evento], format='%H:%M:%S', errors='coerce').dt.time
    dt_entrada = pd.to_datetime(df_final['Data'].astype(str) + ' ' + df_final['Entrada'].astype(str), errors='coerce')
    dt_saida = pd.to_datetime(df_final['Data'].astype(str) + ' ' + df_final['Saída'].astype(str), errors='coerce')
    df_final['Total Horas Trabalhadas'] = (dt_saida - dt_entrada).apply(_formatar_timedelta)
    colunas = ['Data', 'Código Forte', 'Nome', 'Empresa', 'CNPJ', 'Entrada', 'Saída', 'Total Horas Trabalhadas', 'Observação']
    for col in colunas:
        if col not in df_final.columns: df_final[col] = 'N/A'
    df_final = df_final[colunas]
    df_final.rename(columns={'Código Forte': 'Código do Funcionário', 'Nome': 'Nome do Funcionário'}, inplace=True)
    df_final['Data'] = pd.to_datetime(df_final['Data']).dt.strftime('%d/%m/%Y')
    return df_final

def gerar_arquivo_excel(df_organizado, df_bruto, nome_empresa, cnpj, data_inicio, data_fim):
    output_buffer = io.BytesIO()
    periodo_str = f"{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}"
    
    with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
        df_organizado.to_excel(writer, sheet_name='Relatório Diário', index=False, startrow=4)
        df_bruto.to_excel(writer, sheet_name='Log de Eventos (Bruto)', index=False)
        workbook = writer.book
        sheet_diario = writer.sheets['Relatório Diário']
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        font_titulo = Font(name='Calibri', size=16, bold=True)
        font_info = Font(name='Calibri', size=12, bold=True)
        alignment_left = Alignment(horizontal='left', vertical='center')
        sheet_diario.merge_cells('A1:D1')
        titulo_cell = sheet_diario['A1']
        titulo_cell.value = "Relatório de Ponto por Período"
        titulo_cell.font = font_titulo
        titulo_cell.alignment = alignment_left
        sheet_diario['A2'] = "Empresa:"
        sheet_diario['B2'] = nome_empresa
        sheet_diario['A2'].font = font_info
        if cnpj:
            sheet_diario['C2'] = "CNPJ:"
            sheet_diario['D2'] = cnpj
            sheet_diario['C2'].font = font_info
        sheet_diario['A3'] = "Período:"
        sheet_diario['B3'] = periodo_str
        sheet_diario['A3'].font = font_info
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for i, column_cells in enumerate(worksheet.columns, 1):
                max_length = 0
                column_letter = get_column_letter(i)
                start_row = 5 if sheet_name == 'Relatório Diário' else 1
                header_cell = worksheet.cell(row=start_row, column=i)
                if header_cell.value:
                    max_length = len(str(header_cell.value))
                for cell in column_cells:
                    if cell.row > start_row:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                adjusted_width = max(max_length, len(str(header_cell.value)) if header_cell.value else 0) + 2
                worksheet.column_dimensions[column_letter].width = adjusted_width
    output_buffer.seek(0)
    return output_buffer

